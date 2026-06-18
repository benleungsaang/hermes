"""
BOM row types (verified / guessed / implicit) and their visual signals.
Use this in write_excel to pick the right background color and prefix.
"""

from openpyxl.styles import PatternFill

# B 2026-06-17 hard rules
ROW_TYPES = {
    'verified': {
        'fill': None,                      # white / no fill
        'prefix': '',                      # no prefix
        'description': 'MTEXT label + spatial size association succeeded',
    },
    'guessed': {
        'fill': PatternFill('solid', fgColor='FCE4D6'),   # light orange
        'prefix': '[?] ',                  # B wants to see this at a glance
        'description': 'Geometry-only, no MTEXT — derived from guess_unlabeled_geometry',
    },
    'implicit': {
        'fill': PatternFill('solid', fgColor='FFF2CC'),   # light yellow
        'prefix': '[隐] ',                 # packaging-line implicit (e.g. cabinet)
        'description': 'Auto-appended for packaging-line completeness; user may delete',
    },
}


def render_name(row, base_name):
    """Apply prefix per row type. B's hard rule: prefix is in the name column, not separate."""
    rtype = 'verified'
    if row.get('guessed'):
        rtype = 'guessed'
    elif row.get('implicit'):
        rtype = 'implicit'
    return ROW_TYPES[rtype]['prefix'] + base_name


def apply_row_fill(cell, row):
    """Apply the row-type background to a single cell. Caller must call once per column."""
    rtype = 'verified'
    if row.get('guessed'):
        rtype = 'guessed'
    elif row.get('implicit'):
        rtype = 'implicit'
    fill = ROW_TYPES[rtype]['fill']
    if fill is not None:
        cell.fill = fill
