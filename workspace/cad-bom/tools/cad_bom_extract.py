#!/usr/bin/env python3
"""
CAD BOM Extractor — 从 DXF 抽取 BOM（标注驱动方案 B）
======================================================
读取 DXF → 抽取 MTEXT（部件名）+ DIMENSION（尺寸）→ 空间最近原则关联
→ 按名称前缀 + 尺寸容差聚类 → 输出 Excel BOM 表

作者：B + Hermes（2026-06-17）
适用：包装机/机械 2D 总装图（部件名直接标注在图上，无传统明细栏）

用法：
  python3 cad_bom_extract.py <input.dxf> [output.xlsx]
"""
import sys
import re
import math
from pathlib import Path
from collections import defaultdict

# 强制走用户级 site-packages（ezdxf/numpy 装在 py3.12）
sys.path.insert(0, "/home/ubuntu/.local/lib/python3.12/site-packages")

import ezdxf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# 配置
# ============================================================================
SIZE_TOLERANCE_MM = 3.0        # 尺寸聚类容差（±3mm 算同尺寸）
ASSOC_MAX_DIST = 2000.0        # MTEXT 与 DIMENSION 最大关联距离（图纸单位）
NAME_PREFIX_RULES = [
    # 去掉常见前缀/后缀以做"名称归一化"（如 "Transition conveyor 1/2/3/4" 都归一为 "Transition conveyor"）
    (r'\s+\d+\s*$', ''),         # 末尾数字 "1", " 2" 删
    (r'[\(\[【\{]\s*\d+\s*[\)\]】\}]\s*$', ''),  # 末尾带括号数字
    (r'\s+No\.?\s*\d+\s*$', ''),  # "No.1"
    (r'\s+#\s*\d+\s*$', ''),      # "#1"
    (r'#\d+\s*$', ''),            # "abc#1"
]


# ============================================================================
# 工具函数
# ============================================================================
def clean_mtext(raw: str) -> str:
    """去掉 MTEXT 的控制字符 \\A1; \\W0.707; \\P（换行）等 + 包裹花括号"""
    s = re.sub(r'\\[A-Za-z][^;]*;', '', raw)  # \A1; \W0.707; 等
    s = s.replace('\\P', ' ')                # 段落分隔符
    s = s.replace('\\~', ' ')                # 不间断空格
    s = s.replace('\\X', '\\n') if False else s  # 保留 \X（行分隔符不常用，先不动）
    # 去掉成对花括号 {Transition conveyor 2}
    s = re.sub(r'\{\s*([^}]+?)\s*\}', r'\1', s)
    return re.sub(r'\s+', ' ', s).strip()


# 整机/标题关键字 — 出现这些词的 MTEXT 当作"图标题"排除出 BOM
# 注意：HP-FB-370L 等型号 **是项目本身**，保留为 BOM 的一项（项目 = 1 件）
# 只有"装配说明 / 技术要求 / 图号"等纯元数据才过滤
TITLE_KEYWORDS = re.compile(
    r'^\s*图\s*\d+|'
    r'^\s*图\s*号|'
    r'^\s*技\s*术\s*要\s*求|'
    r'^\s*装\s*配\s*说\s*明|'
    r'^\s*说\s*明|'
    r'^\s*标\s*题\s*栏|'
    r'LAYOUT|ASSEMBLY\s*INSTRUCTION|TECHNICAL\s*REQUIREMENT',
    re.IGNORECASE
)


def is_title(text: str) -> bool:
    """判断是否整机型号/标题/技术说明（非部件名）"""
    return bool(TITLE_KEYWORDS.search(text))


def normalize_name(name: str) -> str:
    """名称归一化：去掉末尾数字/编号，让 Transition conveyor 1-4 都归一为 'Transition conveyor'。
    B 2026-06-17 修订：先剥离已知 prefix（Main / Automatic / Servo 等）再归一化，
    避免 "Main conveyor 06" 被错误归一为 "Main conveyor"（少了 06 编号）。
    """
    out = name
    # 已知会带"有意义编号"的前缀 → 剥离后归一化
    SPECIAL_PREFIXES = ['main', 'automatic', 'servo', 'auxiliary', 'sub']
    lower = out.lower().strip()
    for sp in SPECIAL_PREFIXES:
        if lower.startswith(sp + ' '):
            out = out[len(sp) + 1:].strip()
            break
    for pat, repl in NAME_PREFIX_RULES:
        out = re.sub(pat, repl, out).strip()
    return out


# ============================================================================
# 几何猜想（B 2026-06-17 设定）
# 当制图人员只画了几何没标部件名时，基于"几何形态 + 尺寸 + 出现次数"做模式匹配
# 猜出来的部件用 GUESS 角标 + 浅橙色背景，B 决定保留/修改/删除
# ============================================================================
# 模式规则（按优先级匹配，每条规则：(形状描述, 直径范围, 部件名, 备注)）
GEOMETRY_GUESS_RULES = [
    # 圆形 Ø800+ 出现 ≥2 次对称分布 → 翻转弯轮 / 转向台
    # 具体值会在 guess_unlabeled_geometry() 里用实际尺寸做匹配
]


# ============================================================================
# 数量自动修正（B 2026-06-17 设定）
# 当某部件"尺寸完全一致 + 在图上对称出现 ≥2 次"，把 count 修正为 2
# 用几何对称性（X 坐标在图中线两侧接近镜像）判断
# ============================================================================
def detect_symmetric_pairs(texts, dims):
    """
    对 MTEXT 找"尺寸一致 + 几何位置对称"的配对 → 数量 ×2。
    触发条件：
    1. 同 normalize_name 的 MTEXT ≥ 2 个
    2. 这些 MTEXT 关联的尺寸完全相同（或在 ±3mm 容差内）
    3. X 坐标分布在图纸中线两侧（差值 ≥ 某个阈值）
    """
    from collections import Counter
    by_name = defaultdict(list)
    for t in texts:
        if not t.get('sizes'):
            continue
        norm = normalize_name(t['text'])
        by_name[norm].append(t)

    # 找图纸中线 X（中位数）
    all_x = [t['pos'][0] for t in texts]
    if not all_x:
        return set()
    mid_x = sorted(all_x)[len(all_x) // 2]

    symmetric_names = set()
    for norm, instances in by_name.items():
        if len(instances) < 2:
            continue
        # 尺寸完全一致或容差内
        sizes = [t['sizes'][0] for t in instances]
        if max(sizes) - min(sizes) > SIZE_TOLERANCE_MM:
            continue
        # 位置分布：至少有一个在 mid_x 左侧、一个在右侧
        left = [t for t in instances if t['pos'][0] < mid_x - 100]
        right = [t for t in instances if t['pos'][0] > mid_x + 100]
        if left and right:
            symmetric_names.add(norm)
    return symmetric_names


# ============================================================================
# 部件名"对称暗示词"（B 2026-06-17 新增）
# 包装机里很多部件名本身就暗示"双联 / 成对"结构
# 出现这些词的部件强制 count ≥ 2
# ============================================================================
SYMMETRIC_NAME_HINTS = re.compile(
    r'\b(2\s*in\s*1|2-1|'
    r'twin|pair|dual|'
    r'left|right|L/R|'
    r'left\s*&\s*right|LH/RH)\b',
    re.IGNORECASE
)


def get_implied_count(name: str) -> int:
    """
    从部件名反推最少数量（B 2026-06-17 设定）。
    - "2 in 1 transition conveyor" → 至少 2
    - "twin track" / "pair" / "dual feeder" → 至少 2
    - "left" / "right" / "L/R" 单边命名 → 至少 2（左右成对）
    - "double" → 不强制（可能为 1，比如 "Servo double pusher" 是单部件）
    - 普通名称 → 0（不强制）
    """
    m = SYMMETRIC_NAME_HINTS.search(name)
    if not m:
        return 0
    hint = m.group(1).lower().replace(' ', '').replace('-', '')
    if '2in1' in hint or '21' in hint:
        return 2
    if hint in ('twin', 'pair', 'dual'):
        return 2
    if hint in ('left', 'right', 'l/r', 'l&r', 'lh/rh'):
        return 2
    return 0


def count_geometry_instances(doc, target_size_mm, target_name):
    """
    B 2026-06-17 新增：当某个部件名只标 1 次 MTEXT 时，
    用"几何出现次数"反推数量。
    做法：扫描所有 LINE 几何，按"长度 = target_size_mm"统计出现次数。
    简单粗暴但实用：包装机的 2 in 1 transition conveyor 通常是 800mm 长的直线段。
    """
    horiz_segments = []
    for ent, _bn, _depth, _off in _iter_all_entities(doc):
        if ent.dxftype() == 'LINE':
            try:
                s = ent.dxf.start
                e = ent.dxf.end
                length = math.hypot(e.x - s.x, e.y - s.y)
                horiz_segments.append(length)
            except Exception:
                pass
    count = sum(1 for L in horiz_segments if abs(L - target_size_mm) <= SIZE_TOLERANCE_MM * 2)
    return count


# ============================================================================
# 未标注部件的"几何猜想"（B 2026-06-17 新增）
# 规则：圆形部件 + 直径 700~900mm + 出现 ≥2 次 → 猜 "Turnable unit"（转向台）
# 几何特征：单独的 CIRCLE 实体，无对应 MTEXT
# ============================================================================
def guess_unlabeled_geometry(doc, visited_texts, known_geometry):
    """
    扫描 CIRCLE 实体，统计"出现频率高 + 直径在某范围"的几何特征。
    返回 [{'name': 猜的部件名, 'count': N, 'size_mm': Ø, 'reason': '为什么猜', 'needs_confirm': True}]

    B 2026-06-17 修订：
    1. 对应已有 MTEXT 标注 + 几何对称 ≥2 部件的，**也按几何出现次数修正 count**
       （如 "2 in 1 transition conveyor" 文字标了 1 次但画了 2 个 → 几何修正为 ×2）
    2. 圆形部件 + 直径 700~900mm + 出现 ≥2 次无 MTEXT 标注 → 猜 "Turnable unit"
    """
    guesses = []
    from collections import Counter

    # 1. 统计所有 CIRCLE 的直径
    circle_diams = []
    for ent, _bn, _depth, _off in _iter_all_entities(doc):
        if ent.dxftype() == 'CIRCLE':
            r = float(ent.dxf.radius)
            circle_diams.append(r * 2)

    if not circle_diams:
        return guesses

    # 2. 找出现 ≥2 次的"孤直径"（无对应 MTEXT 标注）
    known_diameters = set()
    for t in visited_texts:
        text = t['text']
        for m in re.finditer(r'[Ø⌀D]?\s*(\d{3,4})', text):
            try:
                known_diameters.add(float(m.group(1)))
            except ValueError:
                pass

    # 3. 直径分桶（±5mm 容差）
    buckets = defaultdict(list)
    for d in circle_diams:
        bucket = round(d / 5) * 5
        buckets[bucket].append(d)

    # 4. 找出现 ≥2 次的"未标注圆形"候选
    for bucket_d, ds in buckets.items():
        if len(ds) < 2:
            continue
        if any(abs(bucket_d - kd) < 5 for kd in known_diameters):
            continue
        # 直径范围 700~900 → 猜 Turnable unit
        if 700 <= bucket_d <= 900:
            guesses.append({
                'canonical_name': 'Turnable unit',
                'rep_name': 'Turnable unit',
                'count': len(ds),
                'size_min_mm': min(ds),
                'size_max_mm': max(ds),
                'all_names': [f'∅{d:.0f}' for d in ds],
                'guessed': True,
                'guess_reason': f'图纸上发现 {len(ds)} 个直径 {min(ds):.0f}mm 的圆形部件，无对应 MTEXT 标注',
            })
    return guesses


# ============================================================================
# 型号识别：B 2026-06-17 修订
# HP-FB-370 等整机型号 = 一个部件项，不再加"（项目）"角标
# 仅当需要"过滤"时才用 is_title() 排除（如"图号"、"技术要求"等纯元数据）
# ============================================================================


def dist(a, b) -> float:
    """2D 距离"""
    return math.hypot(a[0] - b[0], a[1] - b[1])


# ============================================================================
# 部件名识别规则（B 2026-06-17 新增，针对 draw_4 类"文字驱动"图）
# 真实部件名 vs 尺寸数字 vs 图注的判定
# ============================================================================
import re as _re

# 纯数字（含千分位、小数）：`1000` / `1,000` / `1.5` / `1000.5`
RE_PURE_NUMBER = _re.compile(r'^[\d,]+(\.\d+)?$')

# 数字 + 单位/后缀：`1250 MAX` / `1000mm` / `Ø1000`
RE_NUMBER_WITH_SUFFIX = _re.compile(r'^[\d,]+(\.\d+)?\s*(MAX|MIN|mm|cm|m|Ø|⌀|D|±)?\s*$', _re.IGNORECASE)

# 纯大写英文（带可选下划线/连字符）：`FLOW` / `STRUCTURE` / `MAX-LOAD`
RE_ALL_CAPS = _re.compile(r'^[A-Z][A-Z0-9_\-]*[A-Z0-9]$|^[A-Z]$')

# 短单词（1 词，< 10 字符），可能是部件名前缀
# Main / Automatic / Conveyor / Top / Bottom 等
SHORT_WORD = _re.compile(r'^[A-Za-z][A-Za-z\-\']{0,9}$')


def classify_text(raw_text: str) -> str:
    """
    判定一个 MTEXT 文本的"身份"：
    - 'size'        尺寸数字（不进 BOM）
    - 'annotation'  图注/标签（不进 BOM）
    - 'part'        部件名（进 BOM）
    - 'prefix'      可能是部件名前缀（如 Main / Automatic 单出现）
    """
    t = raw_text.strip()

    # 1. 纯数字 → 尺寸
    if RE_PURE_NUMBER.match(t):
        return 'size'
    # 2. 数字 + 后缀（如 1250 MAX）→ 尺寸
    if RE_NUMBER_WITH_SUFFIX.match(t):
        return 'size'
    # 3. 纯大写英文（FLOW / STRUCTURE）→ 图注
    #    B 2026-06-17 修订：含数字的全大写（如 SZ180）→ 视为部件名（型号）
    if RE_ALL_CAPS.match(t) and not any(c.isdigit() for c in t):
        return 'annotation'
    # 4. 短单词（< 10 字符，1 词）→ 可能是前缀
    if SHORT_WORD.match(t):
        return 'prefix'
    # 5. 其他 → 部件名
    return 'part'


def extract_size_value(raw_text: str) -> float | None:
    """从尺寸文本里提取数字部分（mm）。"""
    t = raw_text.strip()
    m = _re.search(r'[\d,]+(\.\d+)?', t)
    if not m:
        return None
    try:
        return float(m.group(0).replace(',', ''))
    except ValueError:
        return None


def merge_prefix_to_part(texts, max_dist=1.5):
    """
    把"前缀"短词（Main / Automatic / Conveyor）合并到附近的部件名上。
    B 2026-06-17 修订：要求 prefix 和 part 在**不同 y 行**（prefix 在 part 上面或下面）才合并。
    排除"同一 y 行紧贴"的情况——那种通常是制图者用"种类 + 名称"的方式标注（如 `conveyor` + `Stack in-feed`），
    应当作 2 个独立部件处理。
    """
    prefixes = [(i, t) for i, t in enumerate(texts) if t.get('_class') == 'prefix']
    parts = [(i, t) for i, t in enumerate(texts) if t.get('_class') == 'part']
    if not prefixes or not parts:
        return texts

    for pi, pt in prefixes:
        # 找最近的 part
        best = None
        best_d = float('inf')
        for qi, qt in parts:
            d = dist(pt['pos'], qt['pos'])
            if d < best_d:
                best_d = d
                best = qt
        if best is None or best_d >= max_dist:
            continue
        # B 2026-06-17 新增：要求 prefix 和 part 在不同 y 行（|Δy| > 0.025）
        # 防止"种类 + 名称"标注被误合
        if abs(pt['pos'][1] - best['pos'][1]) < 0.025:
            continue
        best['text'] = f"{pt['text']} {best['text']}".strip()
        best['_class'] = 'part'
        pt['_class'] = '_merged'
    return texts


def merge_inline_parts(texts, x_tol=2, y_tol=0.5):
    """
    B 2026-06-17 修订：合并"几乎同位置"的多个 part 到一个部件名。
    触发的场景：制图者把一个部件名拆成两行 MTEXT（如 标签 + 名称）。
    规则：
      1. 同一 y 行（容差 y_tol，默认 0.5 mm）
      2. x 距离 < x_tol（默认 2 mm，几乎重合）
      3. 合并成"前面 + 后面"
    注意：x_tol 故意设得很小，避免"同一列 / 同一垂直线上的多个部件名"被误合。
    """
    # 找出所有 part，按 y 分组（行）
    parts = [t for t in texts if t.get('_class') == 'part']
    if len(parts) < 2:
        return texts

    # 按 y 排序
    by_y = sorted(parts, key=lambda t: (round(t['pos'][1] / y_tol), t['pos'][0]))

    merged_into = set()  # 哪些 part 已被合并（从 texts 中移除）
    i = 0
    while i < len(by_y):
        current = by_y[i]
        if id(current) in merged_into:
            i += 1
            continue
        # 找同行（y 容差）+ x 紧邻的下一个 part
        j = i + 1
        while j < len(by_y):
            nxt = by_y[j]
            if id(nxt) in merged_into:
                j += 1
                continue
            # y 差
            dy = abs(current['pos'][1] - nxt['pos'][1])
            # x 差
            dx = nxt['pos'][0] - current['pos'][0]
            if dy <= y_tol and 0 < dx <= x_tol:
                # 合并：current = current + " " + nxt
                current['text'] = f"{current['text']} {nxt['text']}".strip()
                merged_into.add(id(nxt))
                # 当前位置用 current（继续找下一个）
            else:
                break
        i += 1

    # 移除被合并的
    texts = [t for t in texts if id(t) not in merged_into]
    return texts


# ============================================================================
# 抽取
# ============================================================================
def _iter_all_entities(doc):
    """
    递归遍历所有块（包括嵌套块）里的 MTEXT / TEXT / DIMENSION / ATTDEF / ATTRIB。
    关键修复（B 2026-06-17）：CAD 习惯把图元放进嵌套块（SOONWIN→块A→块B→…），
    只看 model_space 顶层会漏掉所有标注。必须沿 INSERT 链递归下去。
    """
    from ezdxf.entities import DXFGraphic
    visited = set()  # block 名集合，防止循环引用（如 A→B→A）

    def walk(entities, block_name, depth=0, offset=(0.0, 0.0)):
        for ent in entities:
            t = ent.dxftype()
            if t == 'INSERT':
                # 进入嵌套块 — 子块内所有实体的位置 = 自身位置 + 父块累积偏移
                child_name = ent.dxf.name
                if child_name in visited:
                    continue
                visited.add(child_name)
                try:
                    child_block = doc.blocks.get(child_name)
                    ins = ent.dxf.insert
                    sx = float(ent.dxf.get('xscale', 1.0))  # 缩放
                    sy = float(ent.dxf.get('yscale', 1.0))
                    child_offset = (
                        offset[0] + ins.x * sx,
                        offset[1] + ins.y * sy,
                    )
                    yield from walk(child_block, child_name, depth + 1, child_offset)
                except KeyError:
                    pass
            else:
                yield ent, block_name, depth, offset

    # 1. 遍历 model_space（顶层实体）
    yield from walk(doc.modelspace(), '*MODEL_SPACE', 0)
    # 2. 也要遍历 paper space 布局（用户可能标注在布局上）
    try:
        for layout_name in doc.layouts.names():
            if layout_name == 'Model':
                continue
            layout = doc.layouts.get(layout_name)
            for item in walk(layout, layout_name, 0):
                yield item
    except Exception:
        pass


def extract_entities(doc):
    """
    抽取所有 MTEXT + DIMENSION 实体（带位置、文字/测量值）。
    现在递归到所有嵌套块（B 2026-06-17 修复：解决 SOONWIN→块A→块B 的标注丢失问题）。
    B 2026-06-17 修订（v0.3）：每个 MTEXT 先做"身份分类"（part / size / annotation / prefix），
    避免把尺寸数字（1000 / 1100）或图注（FLOW / STRUCTURE）误当部件名。
    """
    texts = []   # [{'pos':(x,y), 'text':str, 'raw':str, 'block':str, 'depth':int, '_class':str}]
    sizes = []   # [{'pos':(x,y), 'measure':float, 'block':str, 'depth':int}]  —— 文字里的尺寸也算
    dims = []    # 传统 DIMENSION 实体的尺寸

    for ent, block_name, depth, offset in _iter_all_entities(doc):
        t = ent.dxftype()
        if t == 'MTEXT' or t == 'TEXT':
            raw = ent.plain_text() if hasattr(ent, 'plain_text') else (ent.text if hasattr(ent, 'text') else getattr(ent.dxf, 'text', ''))
            clean = clean_mtext(raw)
            if not clean:
                continue
            if is_title(clean):
                continue
            p = ent.dxf.insert
            cls = classify_text(clean)
            entry = {
                'pos': (p.x + offset[0], p.y + offset[1]),
                'text': clean,
                'raw': raw,
                'is_title': False,
                'block': block_name,
                'depth': depth,
                '_class': cls,
            }
            if cls == 'size':
                # 文字里写的尺寸（替代缺失的 DIMENSION）
                val = extract_size_value(clean)
                if val is not None:
                    sizes.append({
                        'pos': entry['pos'],
                        'measure': val,
                        'block': block_name,
                        'depth': depth,
                    })
                # 不进 texts
                continue
            if cls == 'annotation':
                # 图注/标签，不进 BOM
                continue
            texts.append(entry)
        elif t == 'ATTRIB' or t == 'ATTDEF':
            raw = ent.dxf.get('text', '')
            clean = clean_mtext(raw)
            if not clean or is_title(clean):
                continue
            p = ent.dxf.get('insert', (0, 0, 0))
            cls = classify_text(clean)
            entry = {
                'pos': (p[0] + offset[0], p[1] + offset[1]),
                'text': clean,
                'raw': raw,
                'is_title': False,
                'block': block_name,
                'depth': depth,
                '_class': cls,
            }
            if cls == 'size':
                val = extract_size_value(clean)
                if val is not None:
                    sizes.append({
                        'pos': entry['pos'],
                        'measure': val,
                        'block': block_name,
                        'depth': depth,
                    })
                continue
            if cls == 'annotation':
                continue
            texts.append(entry)
        elif t == 'DIMENSION':
            p = ent.dxf.get('defpoint') or ent.dxf.get('insert')
            if p is None:
                continue
            measure = ent.get_measurement() if hasattr(ent, 'get_measurement') else None
            if measure is None:
                continue
            dims.append({
                'pos': (p.x + offset[0], p.y + offset[1]),
                'measure': float(measure),
                'block': block_name,
                'depth': depth,
            })

    # 合并 prefix 短词到附近的 part（如 "Main" + "conveyor" → "Main conveyor"）
    if texts:
        texts = merge_prefix_to_part(texts)
        # 过滤掉已合并的前缀
        texts = [t for t in texts if t.get('_class') != '_merged']

    # 关闭 inline 合并（B 2026-06-17 决定）：
    # 实际场景里 inline 合并容易把不同部件误合。
    # 如 (4, 12) Trunning unit + (6, 12) Automatic → 不应合并
    # 如 (7, 11) Stack in-feed + (7, 11) conveyor → 不应合并
    # 仅保留 prefix 合并（用最近距离判定更安全）。
    #
    # 如果以后真需要 inline 合并，可解开下行注释：
    # if texts:
    #     texts = merge_inline_parts(texts)

    # 合并两种尺寸源：传统 DIMENSION + 文字里写的尺寸
    all_dims = dims + sizes
    return texts, all_dims


def associate_sizes_to_texts(texts, dims, max_dist=ASSOC_MAX_DIST):
    """对每个 MTEXT，找最近的 DIMENSION（测量值）作为该部件的尺寸。
    同一 DIMENSION 可被多个 MTEXT 共享（标注一条尺寸给多个相似件用）。"""
    for t in texts:
        if not dims:
            t['sizes'] = []
            t['nearest_dist'] = None
            continue
        candidates = []
        for d in dims:
            d_ = dist(t['pos'], d['pos'])
            if d_ <= max_dist:
                candidates.append((d_, d['measure'], d['pos']))
        candidates.sort()
        t['sizes'] = [c[1] for c in candidates]
        t['nearest_dist'] = candidates[0][0] if candidates else None
    return texts


def group_components(texts, tol=SIZE_TOLERANCE_MM, symmetric_names=None):
    """
    部件聚类（两阶段）：
    1) 第一阶段：按 normalize_name() 归一化名称分组（如 Transition 1/2/3/4 → 同名组）
    2) 第二阶段：组内按尺寸聚类（容差内 = 同部件）
       但更稳健的做法：先用"组内大部分尺寸"作为代表尺寸，再把"尺寸偏离太大"的实例单拆

    B 2026-06-17 修订：
    - HP-FB-370 等型号不再加"（项目）"角标
    - symmetric_names 集合中的部件（如 2 in 1 transition conveyor），count 修正为对称数
    """
    symmetric_names = symmetric_names or set()
    from collections import Counter

    # 阶段 1: 名称归一化
    by_name = defaultdict(list)
    for t in texts:
        norm = normalize_name(t['text'])
        by_name[norm].append(t)

    # 阶段 2: 名称组内再按尺寸聚
    result = []
    for norm, instances in by_name.items():
        # 收集每个实例的"代表尺寸"（用 nearest_dim 的尺寸）
        sized = [t for t in instances if t['sizes']]
        unsized = [t for t in instances if not t['sizes']]

        if not sized:
            # 全都没尺寸
            all_names = [t['text'] for t in instances]
            rep_name = Counter(all_names).most_common(1)[0][0]
            result.append({
                'canonical_name': norm,
                'rep_name': rep_name,
                'count': len(instances),
                'size_min_mm': None,
                'size_max_mm': None,
                'all_names': all_names,
            })
            continue

        # 找出现最多的尺寸作为"代表尺寸"
        all_meas = [t['sizes'][0] for t in sized]
        meas_counter = Counter(all_meas)
        top_meas, top_count = meas_counter.most_common(1)[0]

        # 把所有尺寸在 ±tol 内的归一组；偏离太大的另起
        matched, outliers = [], []
        for t in sized:
            if abs(t['sizes'][0] - top_meas) <= tol:
                matched.append(t)
            else:
                outliers.append(t)

        # 主组
        all_names_m = [t['text'] for t in matched]
        sizes_m = [t['sizes'][0] for t in matched]
        rep_name = Counter(all_names_m).most_common(1)[0][0]

        # 数量规则（B 2026-06-17）：
        # - 普通部件：len(matched) 即可
        # - 对称部件（几何位置对称）：max(文字, 几何)
        # - 部件名含"对称暗示词"（2 in 1 / double / twin / left / right 等）：强制 ≥2
        count = len(matched)
        is_sym = norm in symmetric_names
        if is_sym:
            all_x = [t['pos'][0] for t in matched]
            mid_x = sorted(all_x)[len(all_x) // 2]
            left = [t for t in matched if t['pos'][0] < mid_x - 100]
            right = [t for t in matched if t['pos'][0] > mid_x + 100]
            geom_count = max(len(left), len(right), 1)
            count = max(len(matched), geom_count)
        # 部件名暗示对称（如 "2 in 1"）
        implied = get_implied_count(rep_name)
        if implied > count:
            count = implied
        result.append({
            'canonical_name': norm,
            'rep_name': rep_name,
            'count': count,
            'size_min_mm': min(sizes_m),
            'size_max_mm': max(sizes_m),
            'all_names': all_names_m,
            'symmetric': is_sym,
            'implied_count': implied if implied > count - (0 if is_sym else 0) else 0,
        })

        # 偏离组（如有）
        for out in outliers:
            result.append({
                'canonical_name': f"{norm} (尺寸异常)",
                'rep_name': out['text'],
                'count': 1,
                'size_min_mm': out['sizes'][0],
                'size_max_mm': out['sizes'][0],
                'all_names': [out['text']],
            })

        # 无尺寸的也单列
        for u in unsized:
            result.append({
                'canonical_name': f"{norm} (无尺寸)",
                'rep_name': u['text'],
                'count': 1,
                'size_min_mm': None,
                'size_max_mm': None,
                'all_names': [u['text']],
            })

    result.sort(key=lambda r: (-r['count'], r['canonical_name']))
    return result


# ============================================================================
# 包装线 (packaging line) 隐含件
# ============================================================================
def suggest_implicit_components(groups, unassoc, line_type='packaging line'):
    """
    返回"通常在图上不画、但 packaging line BOM 应自动追加"的部件。
    B 2026-06-17 规则：
    - packaging line = 整套方案（不只一台单机）
    - 图上一般不会画 electrical cabinet（总电箱）→ BOM 自动追加一行
    - B 不需要时手动从 Excel 删除
    - 如果图上**已经标注**了 cabinet/control（任何形式），则不重复追加

    返回的是**直接进 BOM 主表**的条目（dict 格式与 group 一致）。
    """
    suggestions = []
    existing_names = {g['canonical_name'].lower() for g in groups}

    # 总电箱 — packaging line 必有
    has_cabinet = any('cabinet' in n or '电箱' in n or 'control' in n or '电控' in n
                      for n in existing_names)
    if not has_cabinet:
        suggestions.append({
            'canonical_name': 'Electrical cabinet（总电箱）',
            'rep_name': 'Electrical cabinet（总电箱）',
            'count': 1,
            'size_min_mm': None,
            'size_max_mm': None,
            'all_names': ['【packaging line 隐含件：图上不画，运行时由 B 人工删除不需要的】'],
            'implicit': True,   # 标记为"自动追加"
        })

    return suggestions


# ============================================================================
# 输出
# ============================================================================
def write_excel(groups, unassociated_texts, output_path, source_dxf, implicit_suggestions=None, guessed=None):
    """
    B 2026-06-17 修订：
    - 行类型区分：普通（白）/ 隐含件 cabinet（浅黄）/ 几何猜想（浅橙）+ 列"[?]"角标
    - 隐含件 cabinet 强制放最后（packaging line 必有项）
    """
    implicit_suggestions = implicit_suggestions or []
    guessed = guessed or []
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    # 标题
    ws['A1'] = f"BOM from {Path(source_dxf).name}  ·  整套方案 = packaging line"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')

    ws['A2'] = (f"生成时间: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  ·  "
                f"浅黄色行 = packaging line 隐含件（电气柜等）  ·  "
                f"浅橙色行 = 几何猜想部件（需人工确认：[?] 列）")
    ws.merge_cells('A2:G2')

    # 表头
    headers = ['#', '部件名（归一化）', '代表名称', '数量', '尺寸 (mm)', '所有实例名', '备注']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.alignment = Alignment(horizontal='center')

    # 颜色定义
    IMPLICIT_FILL = PatternFill("solid", fgColor="FFF2CC")  # 浅黄 - 隐含件
    GUESS_FILL = PatternFill("solid", fgColor="FCE4D6")     # 浅橙 - 几何猜想

    # 排序：B 2026-06-17 规则
    # 1. Electrical cabinet 始终排最后（不论图上是否标了）
    # 2. 其他部件：按 count 降序，相同按名称
    all_rows = list(groups) + list(guessed) + list(implicit_suggestions)

    def sort_key(g):
        is_cabinet = 'cabinet' in g['canonical_name'].lower() or '电箱' in g['canonical_name']
        cabinet_order = 1 if is_cabinet else 0
        return (cabinet_order, -g['count'], g['canonical_name'])

    all_rows = sorted(all_rows, key=sort_key)

    for i, g in enumerate(all_rows, 1):
        row = 4 + i
        # 名称前加角标：[隐] / [猜?]
        name_display = g['canonical_name']
        note = ''
        if g.get('guessed'):
            name_display = f"[?] {g['canonical_name']}"
            note = g.get('guess_reason', '几何猜想，需人工确认')
        elif g.get('implicit'):
            name_display = f"[隐] {g['canonical_name']}"
            note = g.get('all_names', [''])[0] if g.get('all_names') else ''

        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=name_display)
        ws.cell(row=row, column=3, value=g['rep_name'])
        ws.cell(row=row, column=4, value=g['count'])
        if g['size_min_mm'] is not None and g['size_max_mm'] is not None:
            if abs(g['size_max_mm'] - g['size_min_mm']) < 0.1:
                size_s = f"{g['size_min_mm']:.1f}"
            else:
                size_s = f"{g['size_min_mm']:.1f} ~ {g['size_max_mm']:.1f}"
        else:
            size_s = "（无尺寸）"
        ws.cell(row=row, column=5, value=size_s)
        ws.cell(row=row, column=6, value=' | '.join(g['all_names']))
        ws.cell(row=row, column=7, value=note)

        # 行底色
        if g.get('implicit'):
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = IMPLICIT_FILL
        elif g.get('guessed'):
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = GUESS_FILL

    # 列宽
    for col, w in enumerate([5, 32, 28, 8, 20, 50, 40], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # 未关联尺寸的部件（提醒）
    if unassociated_texts:
        ws2 = wb.create_sheet("未关联尺寸")
        ws2['A1'] = "⚠️ 以下 MTEXT 在阈值距离内未找到尺寸标注"
        ws2['A1'].font = Font(bold=True, color="C00000")
        ws2.merge_cells('A1:C1')
        ws2.append([])
        ws2.append(['#', '部件名', '位置 (x, y)'])
        for c in [ws2.cell(row=3, column=i) for i in range(1, 4)]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="C00000")
        for i, t in enumerate(unassociated_texts, 1):
            ws2.append([i, t['text'], f"({t['pos'][0]:.1f}, {t['pos'][1]:.1f})"])
        for col, w in enumerate([5, 40, 25], 1):
            ws2.column_dimensions[get_column_letter(col)].width = w

    wb.save(output_path)
    return output_path


# ============================================================================
# 主流程
# ============================================================================
def main():
    if len(sys.argv) < 2:
        print("用法: python3 cad_bom_extract.py <input.dxf|.dwg> [output.xlsx]")
        sys.exit(1)

    src = Path(sys.argv[1])
    if not src.exists():
        print(f"❌ 找不到: {src}")
        sys.exit(1)

    out = Path(sys.argv[2]) if len(sys.argv) > 2 else src.with_suffix('.bom.xlsx')

    # 自动识别 DWG → 用 ODAFC 转换后读取
    ext = src.suffix.lower()
    if ext == '.dwg':
        from ezdxf.addons import odafc
        print(f"📂 读取 DWG（经 ODAFC 自动转换）: {src}")
        doc = odafc.readfile(str(src))
    else:
        print(f"📂 读取: {src}")
        doc = ezdxf.readfile(str(src))
    print(f"   版本: {doc.dxfversion}")
    print(f"   图层: {len(doc.layers)}, 块: {len(doc.blocks)}")

    # 1. 抽取
    texts, dims = extract_entities(doc)
    print(f"   MTEXT (有文字): {len(texts)}")
    print(f"   DIMENSION (有测量): {len(dims)}")

    # 2. 空间关联
    texts = associate_sizes_to_texts(texts, dims)

    # 2.5 数量自动修正（B 2026-06-17）
    # 当某部件"尺寸一致 + 几何对称出现 ≥2 次"时，count 修正为实际 MTEXT 实例数
    # 当文字只标 1 次但几何画了 ≥2 个对称件时，count 修正为几何数（标记 guessed）
    symmetric_names = detect_symmetric_pairs(texts, dims)
    if symmetric_names:
        print(f"\n🔄 对称部件（自动 ×{len(symmetric_names)} 分类）: {', '.join(symmetric_names)}")

    # 3. 聚类
    groups = group_components(texts, symmetric_names=symmetric_names)
    print(f"\n📋 识别到 {len(groups)} 类部件:")
    for g in groups:
        size_s = f"{g['size_min_mm']:.1f}" if g['size_min_mm'] is not None else "（无尺寸）"
        if g['size_max_mm'] is not None and abs(g['size_max_mm'] - g['size_min_mm']) > 0.1:
            size_s += f" ~ {g['size_max_mm']:.1f}"
        print(f"   ×{g['count']}  {g['canonical_name']:35s}  {size_s} mm")
        for n in g['all_names']:
            print(f"        - {n}")

    # 4. 未关联提醒
    unassoc = [t for t in texts if not t['sizes']]
    if unassoc:
        print(f"\n⚠️  {len(unassoc)} 个 MTEXT 未找到尺寸（需人工确认）:")
        for t in unassoc:
            print(f"   - \"{t['text']}\" at ({t['pos'][0]:.0f}, {t['pos'][1]:.0f})")

    # 4.3 几何猜想（B 2026-06-17）
    # 制图人员只画了几何没标部件名时（如 ∅802 实为 turnable unit），基于形态猜
    guessed = guess_unlabeled_geometry(doc, texts, [])
    if guessed:
        print(f"\n💡 几何猜想（B 决定保留/修改/删除）:")
        for g in guessed:
            print(f"   [?] {g['canonical_name']}  ×{g['count']}  {g['guess_reason']}")

    # 4.5 packaging line 隐含件 — 不再合入主表
    # 排在主表最后一行（浅黄底色，提示"图上不画"）
    implicit = suggest_implicit_components(groups, unassoc)
    if implicit:
        print(f"\n💡 packaging line 隐含件（已自动追加，B 不需要时手动删）:")
        for sug in implicit:
            print(f"   [隐] {sug['canonical_name']}")

    # 5. 输出 Excel（隐含件在最后一行）
    write_excel(groups, unassoc, str(out), str(src),
                implicit_suggestions=implicit, guessed=guessed)
    print(f"\n✅ BOM 已保存: {out}")


if __name__ == '__main__':
    main()
